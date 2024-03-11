import tkinter as tk
from tkinter import PhotoImage, ttk
from PIL import Image, ImageTk
import customtkinter as ctk
import pyodbc
import os
import re
import sys
import configparser
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font

#####################################################################################################################################################

#Configparser
config = configparser.ConfigParser()
config.read(os.path.abspath("scripts\Exportacion_Manual\Config.ini"))

#####################################################################################################################################################

#Geometry
ScreenHeight = int(config["GEOMETRIA"]["ALTO"])
ScreenWidth = int(config["GEOMETRIA"]["ANCHO"])

#####################################################################################################################################################

# Conexión a la base de datos
server = str(config["DB"]["DBSERVER"]) 
database = str(config["DB"]["DBNAME"]) 
conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';Trusted_Connection=yes')

#####################################################################################################################################################

def export_Info(tabla, desde, hasta, type):
    cursor = conn.cursor()
    try:
        if type=="Reports":
            query = f"SELECT Reportes_Texto_Reporte, Reportes_Hora_Reporte, Reportes_Numero_Reporte, Reportes_Hora_Real_Reporte, Reportes_Numero_Real_Reporte FROM {tabla} WHERE Time_Stamp >= '{desde}' AND Time_Stamp <= '{hasta} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["REPORTES"])
            encabezados = ('Texto Del Reporte', 'Hora Del Reporte', 'Numero Del Reporte', 'Hora Real Del Reporte', 'Numero Real Del Reporte')
            CantidadValores = 5
        elif type =="Exits":
            query = f"SELECT Now_Local, sys_On_Off FROM {tabla} WHERE Time_Stamp >= '{desde}' AND Time_Stamp <= '{hasta} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["SALIDAS"])
            encabezados = ['Hora De Accion', 'Tipo De Accion']
            CantidadValores = 2
        elif type=="Habs":
            NumeroHab = re.findall(r'\d+', tabla)
            query = f"SELECT Time_Stamp, Hab_{int(NumeroHab[0])-1}_Tiempo_Limpiando FROM {tabla} WHERE Time_Stamp >= '{desde}' AND Time_Stamp <= '{hasta} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["HABITACIONES"])
            encabezados = ['Hora Terminada', f'Tiempo Limpiando Habitación {int(NumeroHab[0])}']
            CantidadValores = 2
        elif type =="Faults":
            query = f"SELECT Time_Stamp, AveriaR_Texto, AveriaR_Numero FROM {tabla} WHERE Time_Stamp >= '{desde}' AND Time_Stamp <= '{hasta} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["AVERIAS"])
            encabezados = ['Hora De Accion', 'Texto Averia', 'Numero Averia']
            CantidadValores = 3
        elif type =="Forces":
            query = f"SELECT Time_Stamp, ForzadoR_Texto, ForzadoR_Numero FROM {tabla} WHERE Time_Stamp >= '{desde}' AND Time_Stamp <= '{hasta} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["FORZADOS"])
            encabezados = ['Hora De Accion', 'Texto Forzado', 'Numero Forzado']
            CantidadValores = 3
        cursor.execute(query)
        rows = cursor.fetchall()
        rows = [(re.sub(r"[\(\)]", "", item) if isinstance(item, str) else item) for row in rows for item in row]
    

        if not os.path.exists(carpeta_exports):
            os.makedirs(carpeta_exports)
    
                
        exportar_excel(rows, encabezados, CantidadValores, carpeta_exports, tabla)
        show_messagebox_ok()

        
    except Exception as e:
        print(e)
        if re.search(r"Errno 13", str(e)):
            time = datetime.now().strftime("%d-%m-%y")
            show_messagebox_nok(f"El archivo {tabla} {time}.xlsx está abierto por otra aplicación", 40)
        if re.search(r"WinError 3", str(e)):
            show_messagebox_nok(f"El sistema no puede encontrar la ruta especificada\nComprueba que la ruta está bien escrita", 30)
        if re.search(r"list index out of range", str(e)):
            show_messagebox_nok(f"Seleccione Una Tabla De Habitación Existente", 40)
#####################################################################################################################################################

def show_messagebox_nok(message, pad):
    root.attributes("-topmost", False)
    ventana_emergente = tk.Toplevel(root)
    label_ventana_emergente = tk.Label(ventana_emergente, text=message, font=("Arial", 12))
    label_ventana_emergente.pack(pady=pad)
    ventana_emergente.overrideredirect(True)
    ventana_emergente.geometry(f"600x100+{int((root.winfo_screenwidth() / 2) - 300)}+{int((root.winfo_screenheight() / 2) - 50)}")
    SysCloseButton = tk.Button(ventana_emergente, image=x_photo, borderwidth=0, command=lambda:destroy_popup_nok(ventana_emergente))
    SysCloseButton.place(relx=1, x=0, y=0, anchor='ne')
    root.after(5000, lambda:destroy_popup_nok(ventana_emergente))

#####################################################################################################################################################   

def destroy_popup_nok(ventana_emergente):
    ventana_emergente.destroy()
    root.attributes("-topmost", True)

#####################################################################################################################################################

def show_messagebox_ok():
    root.attributes("-topmost", False)
    GoBack(False)
    ventana_emergente = tk.Toplevel(root)
    label_ventana_emergente = tk.Label(ventana_emergente, text="Exportación Correcta", font=("Arial", 12))
    label_ventana_emergente.pack(pady=40)
    ventana_emergente.overrideredirect(True)
    ventana_emergente.geometry(f"200x100+{int((root.winfo_screenwidth() / 2) - 100)}+{int((root.winfo_screenheight() / 2) - 50)}")
    root.after(2000, lambda:destroy_popup_ok(ventana_emergente))

#####################################################################################################################################################   

def destroy_popup_ok(ventana_emergente):
    ventana_emergente.destroy()
    root.attributes("-topmost", True)
    FrameExportaciones.place(relx=0.5, rely=0.5, anchor='center')
    
#####################################################################################################################################################

def exportar_excel(rows, encabezados,CantidadValores, carpeta_exports, tabla):
    wb = openpyxl.Workbook()
    sheet = wb.active
    #Encabezado
    for col, encabezado in enumerate(encabezados, 1):
        sheet.cell(row=1, column=col, value=encabezado)
        sheet.cell(row=1, column=col).font = Font(bold=True)
    #Lineas
    x=0
    y=0
    for col, rowData in enumerate(rows, 1):
        y=y+1
        sheet.cell(row=x+2, column=y, value=rowData)
        if col % CantidadValores == 0:
            x=x+1
            y=0

    wb.save(f'{carpeta_exports}\{tabla} {datetime.now().strftime("%d-%m-%y")}.xlsx')

#####################################################################################################################################################

def set_actual_date(type):
    ActualDate= datetime.now().strftime('%Y-%m-%d')
    if type=="Reports":
        ToEntryReports.delete(0, 'end')
        ToEntryReports.insert(0, ActualDate)
    elif type=="Exits":
        ToEntryExits.delete(0, 'end')
        ToEntryExits.insert(0, ActualDate)
    elif type=="Habs":
        ToEntryHabs.delete(0, 'end')
        ToEntryHabs.insert(0, ActualDate)
    elif type=="Faults":
        ToEntryHabs.delete(0, 'end')
        ToEntryHabs.insert(0, ActualDate)
    elif type=="Forces":
        ToEntryHabs.delete(0, 'end')
        ToEntryHabs.insert(0, ActualDate)

#####################################################################################################################################################

def export_time(time, type):
    ar_time = []
    yesterday = datetime.now() - timedelta(days=1)
    ar_time.append(yesterday.strftime('%Y-%m-%d'))
    ar_time.append(date.today().replace(day=1).strftime('%Y-%m-%d'))
    ar_time.append(datetime.now().strftime('%Y-%m-%d'))
    if type=="Reports":
        FromEntryReports.delete(0, 'end')
        FromEntryReports.insert(0, ar_time[time])
        ToEntryReports.delete(0, 'end')
        ToEntryReports.insert(0, ar_time[2])
        export(type)
    elif type=="Exits":
        FromEntryExits.delete(0, 'end')
        FromEntryExits.insert(0, ar_time[time])
        ToEntryExits.delete(0, 'end')
        ToEntryExits.insert(0, ar_time[2])
        export(type)
    elif type=="Habs":
        FromEntryHabs.delete(0, 'end')
        FromEntryHabs.insert(0, ar_time[time])
        ToEntryHabs.delete(0, 'end')
        ToEntryHabs.insert(0, ar_time[2])
        export(type)
    elif type=="Faults":
        FromEntryFaults.delete(0, 'end')
        FromEntryFaults.insert(0, ar_time[time])
        ToEntryFaults.delete(0, 'end')
        ToEntryFaults.insert(0, ar_time[2])
        export(type)
    elif type=="Forces":
        FromEntryForces.delete(0, 'end')
        FromEntryForces.insert(0, ar_time[time])
        ToEntryForces.delete(0, 'end')
        ToEntryForces.insert(0, ar_time[2])
        export(type)
    

#####################################################################################################################################################    

def export(type):
    if type=="Reports":
        table = "Reportes"
        since = FromEntryReports.get()
        till = ToEntryReports.get()
    elif type=="Exits":
        table = "Salida_De_Programa"
        since = FromEntryExits.get()
        till = ToEntryExits.get()
    elif type=="Habs":
        table = TableComboxHabs.get()
        since = FromEntryHabs.get()
        till = ToEntryHabs.get()
    elif type=="Forces":
        table = "Justificar_Forzado"
        since = FromEntryForces.get()
        till = ToEntryForces.get()
    elif type=="Faults":
        table = "Justificar_Averia"
        since = FromEntryFaults.get()
        till = ToEntryFaults.get()
    
    
    export_Info(table, since, till, type)

#####################################################################################################################################################

def GoBack(all):
    GoBackButton.place_forget()
    FrameExportReports.place_forget()
    FrameExportExit.place_forget()
    FrameExportHabs.place_forget()
    FrameExportForces.place_forget()
    FrameExportFaults.place_forget()
    if all:
        FrameExportaciones.place(relx=0.5, rely=0.5, anchor='center')

#####################################################################################################################################################

def show_Reports():
    FrameExportaciones.place_forget()
    FrameExportReports.place(relx=0.5, rely=0.5, anchor='center')
    GoBackButton.place(relx=0, x=0, y=0, anchor='nw')

def show_Exit():
    FrameExportaciones.place_forget()
    FrameExportExit.place(relx=0.5, rely=0.5, anchor='center')
    GoBackButton.place(relx=0, x=0, y=0, anchor='nw')

def show_Habs():
    FrameExportaciones.place_forget()
    FrameExportHabs.place(relx=0.5, rely=0.5, anchor='center')
    GoBackButton.place(relx=0, x=0, y=0, anchor='nw')

def show_Faults():
    FrameExportaciones.place_forget()
    FrameExportFaults.place(relx=0.5, rely=0.5, anchor='center')
    GoBackButton.place(relx=0, x=0, y=0, anchor='nw')

def show_Forces():
    FrameExportaciones.place_forget()
    FrameExportForces.place(relx=0.5, rely=0.5, anchor='center')
    GoBackButton.place(relx=0, x=0, y=0, anchor='nw')

#####################################################################################################################################################

# Interfaz de usuario
root = tk.Tk()
root.title("Exportar Información")
root.config(bg='#505050')

#####################################################################################################################################################

#Frames
FrameExportaciones = tk.Frame(root, width=int(ScreenWidth/2), height=int(ScreenHeight/2), bg= '#505050')
FrameExportReports = tk.Frame(root, bg= '#505050')
FrameExportExit = tk.Frame(root, bg= '#505050')
FrameExportHabs = tk.Frame(root, bg= '#505050')
FrameExportForces = tk.Frame(root, bg= '#505050')
FrameExportFaults = tk.Frame(root, bg= '#505050')

#####################################################################################################################################################

#Scree Gometry
ScreeGometry = f'{ScreenWidth}x{ScreenHeight}+{int((root.winfo_screenwidth() / 2) - (ScreenWidth / 2))}+{int((root.winfo_screenheight() / 2) - (ScreenHeight / 2))}'
root.geometry(ScreeGometry)
root.overrideredirect(True)
root.attributes("-topmost", True)

#####################################################################################################################################################

#Close App
x_photo = PhotoImage(file=os.path.abspath("scripts\\Exportacion_Manual\\x_button.png"))
SysCloseButton = tk.Button(root, image=x_photo, borderwidth=0, bg= '#505050', command=sys.exit)
SysCloseButton.place(relx=1, x=0, y=0, anchor='ne')

#Go Back
back_phooto = PhotoImage(file=os.path.abspath("scripts\\Exportacion_Manual\\Back_Button.png"))
GoBackButton = tk.Button(root, image=back_phooto, borderwidth=0, bg= '#505050', command=lambda:GoBack(True))

#####################################################################################################################################################

#Initial Frame
FrameExportaciones.place(relx=0.5, rely=0.5, anchor='center')

#####################################################################################################################################################

#Open Reports
OpenFrameReportes = ctk.CTkButton(FrameExportaciones, text="Exportar Reportes", command=show_Reports)
OpenFrameReportes.grid(row=1, column=0, padx=10, pady=10)

#To Frame Reports
LabelReports = tk.Label(FrameExportReports, text="Exportador Tabla Reportes", bg= '#505050', fg='white')
LabelReports.grid(row=0, columnspan=3, pady=5)

FromLabelReports = tk.Label(FrameExportReports, text="Desde:", bg= '#505050', fg='white')
FromLabelReports.grid(row=1, column=0)
FromEntryReports = tk.Entry(FrameExportReports, width=12, font= 10)
FromEntryReports.grid(row=1, column=1)

ToLabelReports = tk.Label(FrameExportReports, text="Hasta:", bg='#505050', fg='white')
ToLabelReports.grid(row=2, column=0)
ToEntryReports = tk.Entry(FrameExportReports, width=12, font= 10)
ToEntryReports.grid(row=2, column=1)

#Aplicate Actual Date
ActualDateButtonReports = ctk.CTkButton(FrameExportReports, text="Fecha actual", border_width=1, command=lambda:set_actual_date(str("Reports")))
ActualDateButtonReports.grid(row=2, column=2, padx=5)

#Export day
ExportLastDayReports = ctk.CTkButton(FrameExportReports, text="Exportar Último Dia", command=lambda:export_time(0, (str("Reports"))))
ExportLastDayReports.grid(row=1, column=4, padx=5, pady=2)

#Export month
ExportLastMonthReports = ctk.CTkButton(FrameExportReports, text="Exportar Último Mes", command=lambda:export_time(1, (str("Reports"))))
ExportLastMonthReports.grid(row=2, column=4, padx=5, pady=2)

#Export
ExportButtonReports = ctk.CTkButton(FrameExportReports, text="Exportar", command=lambda:export(str("Reports")))
ExportButtonReports.grid(row=3, column=1, columnspan=1, pady=(10, 0), sticky="nsew")

#####################################################################################################################################################

#Open Exits
OpenFrameExit = ctk.CTkButton(FrameExportaciones, text="Exportar Salidas Del Sistema", command=show_Exit)
OpenFrameExit.grid(row=1, column=1, padx=10, pady=10)

#To Frame Exits
LabelExits = tk.Label(FrameExportExit, text="Exportador Tabla Reportes", bg= '#505050', fg='white')
LabelExits.grid(row=0, columnspan=3, pady=5)

FromLabelExits = tk.Label(FrameExportExit, text="Desde:", bg= '#505050', fg='white')
FromLabelExits.grid(row=1, column=0)
FromEntryExits = tk.Entry(FrameExportExit, width=12, font= 10)
FromEntryExits.grid(row=1, column=1)

ToLabelExits = tk.Label(FrameExportExit, text="Hasta:", bg='#505050', fg='white')
ToLabelExits.grid(row=2, column=0)
ToEntryExits = tk.Entry(FrameExportExit, width=12, font= 10)
ToEntryExits.grid(row=2, column=1)

#Aplicate Actual Date
ActualDateButtonExits = ctk.CTkButton(FrameExportExit, text="Fecha actual", border_width=1, command=lambda:set_actual_date(str("Exits")))
ActualDateButtonExits.grid(row=2, column=2)

#Export day
ExportLastDayExits = ctk.CTkButton(FrameExportExit, text="Exportar Último Dia", command=lambda:export_time(0, (str("Exits"))))
ExportLastDayExits.grid(row=1, column=4, padx=5, pady=2)

#Export month
ExportLastMonthExits = ctk.CTkButton(FrameExportExit, text="Exportar Último Mes", command=lambda:export_time(1, (str("Exits"))))
ExportLastMonthExits.grid(row=2, column=4, padx=5, pady=2)

#Export
ExportButtonExits = ctk.CTkButton(FrameExportExit, text="Exportar", command=lambda:export(str("Exits")))
ExportButtonExits.grid(row=3, column=1, columnspan=1, pady=(10, 0), sticky="nsew")

#####################################################################################################################################################

#Open Habs
OpenFrameHabs = ctk.CTkButton(FrameExportaciones, text="Exportar Limpieza Habitaciones", command=show_Habs)
OpenFrameHabs.grid(row=1, column=2, padx=10, pady=10)

#To Frame Habs
LabelHabs = tk.Label(FrameExportHabs, text="Exportador Tabla Habitaciones", bg= '#505050', fg='white')
LabelHabs.grid(row=0, columnspan=6, pady=5)

TableHabs = tk.Label(FrameExportHabs, text="Tabla:", bg= '#505050', fg='white')
TableHabs.grid(row=1, column=1)

with open('scripts\\Exportacion_Manual\\archivo_valores.txt', 'r') as file:
    ListHabs = file.readlines()
    ListHabs = [valor.strip() for valor in ListHabs]
TableComboxHabs = ttk.Combobox(FrameExportHabs, values=ListHabs, width=31)
TableComboxHabs.grid(row=1, column=2, padx=2)

FromLabelHabs = tk.Label(FrameExportHabs, text="Desde:", bg= '#505050', fg='white')
FromLabelHabs.grid(row=1, column=3)
FromEntryHabs = tk.Entry(FrameExportHabs)
FromEntryHabs.grid(row=1, column=4)

ToLabelHabs = tk.Label(FrameExportHabs, text="Hasta:", bg= '#505050', fg='white')
ToLabelHabs.grid(row=2, column=3)
ToEntryHabs = tk.Entry(FrameExportHabs)
ToEntryHabs.grid(row=2, column=4)

#Aplicate Actual Date
ActualDateButtonHabs = ctk.CTkButton(FrameExportHabs, text="Fecha actual", border_width=1, command=lambda:set_actual_date(str("Habs")))
ActualDateButtonHabs.grid(row=2, column=2)

#Export day
ExportLastDayHabs = ctk.CTkButton(FrameExportHabs, text="Exportar Último Dia", command=lambda:export_time(0, (str("Habs"))))
ExportLastDayHabs.grid(row=1, column=5, padx=5, pady=2)

#Export month
ExportLastMonthHabs = ctk.CTkButton(FrameExportHabs, text="Exportar Último Mes", command=lambda:export_time(1, (str("Habs"))))
ExportLastMonthHabs.grid(row=2, column=5, padx=5, pady=2)

#Export
ExportButtonHabs = ctk.CTkButton(FrameExportHabs, text="Exportar", command=lambda:export(str("Habs")))
ExportButtonHabs.grid(row=3, column=4, columnspan=1, pady=(10, 0), sticky="nsew")

#####################################################################################################################################################

#Open Faults
OpenFrameFaults = ctk.CTkButton(FrameExportaciones, text="Exportar Averias", command=show_Faults)
OpenFrameFaults.grid(row=1, column=3, padx=10, pady=10)

#To Frame Faults
LabelFaults = tk.Label(FrameExportFaults, text="Exportador Tabla Averias", bg= '#505050', fg='white')
LabelFaults.grid(row=0, columnspan=3, pady=5)

FromLabelFaults = tk.Label(FrameExportFaults, text="Desde:", bg= '#505050', fg='white')
FromLabelFaults.grid(row=1, column=0)
FromEntryFaults = tk.Entry(FrameExportFaults, width=12, font= 10)
FromEntryFaults.grid(row=1, column=1)

ToLabelFaults = tk.Label(FrameExportFaults, text="Hasta:", bg='#505050', fg='white')
ToLabelFaults.grid(row=2, column=0)
ToEntryFaults = tk.Entry(FrameExportFaults, width=12, font= 10)
ToEntryFaults.grid(row=2, column=1)

#Aplicate Actual Date
ActualDateButtonFaults = ctk.CTkButton(FrameExportFaults, text="Fecha actual", border_width=1, command=lambda:set_actual_date(str("Faults")))
ActualDateButtonFaults.grid(row=2, column=2, padx=5)

#Export day
ExportLastDayFaults = ctk.CTkButton(FrameExportFaults, text="Exportar Último Dia", command=lambda:export_time(0, (str("Faults"))))
ExportLastDayFaults.grid(row=1, column=4, padx=5, pady=2)

#Export month
ExportLastMonthFaults = ctk.CTkButton(FrameExportFaults, text="Exportar Último Mes", command=lambda:export_time(1, (str("Faults"))))
ExportLastMonthFaults.grid(row=2, column=4, padx=5, pady=2)

#Export
ExportButtonFaults = ctk.CTkButton(FrameExportFaults, text="Exportar", command=lambda:export(str("Faults")))
ExportButtonFaults.grid(row=3, column=1, columnspan=1, pady=(10, 0), sticky="nsew")

#####################################################################################################################################################

#Open Forces
OpenFrameForces = ctk.CTkButton(FrameExportaciones, text="Exportar Averias", command=show_Forces)
OpenFrameForces.grid(row=1, column=4, padx=10, pady=10)

#To Frame Forces
LabelForces = tk.Label(FrameExportForces, text="Exportador Tabla Averias", bg= '#505050', fg='white')
LabelForces.grid(row=0, columnspan=3, pady=5)

FromLabelForces = tk.Label(FrameExportForces, text="Desde:", bg= '#505050', fg='white')
FromLabelForces.grid(row=1, column=0)
FromEntryForces = tk.Entry(FrameExportForces, width=12, font= 10)
FromEntryForces.grid(row=1, column=1)

ToLabelForces = tk.Label(FrameExportForces, text="Hasta:", bg='#505050', fg='white')
ToLabelForces.grid(row=2, column=0)
ToEntryForces = tk.Entry(FrameExportForces, width=12, font= 10)
ToEntryForces.grid(row=2, column=1)

#Aplicate Actual Date
ActualDateButtonForces = ctk.CTkButton(FrameExportForces, text="Fecha actual", border_width=1, command=lambda:set_actual_date(str("Forces")))
ActualDateButtonForces.grid(row=2, column=2, padx=5)

#Export day
ExportLastDayForces = ctk.CTkButton(FrameExportForces, text="Exportar Último Dia", command=lambda:export_time(0, (str("Forces"))))
ExportLastDayForces.grid(row=1, column=4, padx=5, pady=2)

#Export month
ExportLastMonthForces = ctk.CTkButton(FrameExportForces, text="Exportar Último Mes", command=lambda:export_time(1, (str("Forces"))))
ExportLastMonthForces.grid(row=2, column=4, padx=5, pady=2)

#Export
ExportButtonForces = ctk.CTkButton(FrameExportForces, text="Exportar", command=lambda:export(str("Forces")))
ExportButtonForces.grid(row=3, column=1, columnspan=1, pady=(10, 0), sticky="nsew")

#####################################################################################################################################################

#Styles 
def adjust_size(width, height):
    title_font = ('Arial', (int(width // 25) - int(height // 50)), 'bold')
    text_font = ('Helvetica', (int(width // 40) - int(height // 80)), 'bold')
    start_button_font = ('Arial', (int(width // 45) - int(height // 55)))
    button_font = ('Arial', (int(width // 50) - int(height // 100)))
    combox_font = ('Helvetica', (int(width // 50) - int(height // 80)))
    OpenFrameReportes.configure(font=start_button_font)
    OpenFrameExit.configure(font=start_button_font)
    OpenFrameHabs.configure(font=start_button_font)
    OpenFrameFaults.configure(font=start_button_font)
    OpenFrameForces.configure(font=start_button_font)
    LabelReports.config(font=title_font)
    ToLabelReports.configure(font=text_font)
    ToEntryReports.configure(font=text_font)
    FromLabelReports.configure(font=text_font)
    FromEntryReports.configure(font=text_font)
    ActualDateButtonReports.configure(font=button_font)
    ExportButtonReports.configure(font=button_font)
    ExportLastDayReports.configure(font=button_font)
    ExportLastMonthReports.configure(font=button_font)
    LabelExits.config(font=title_font)
    ToLabelExits.configure(font=text_font)
    ToEntryExits.configure(font=text_font)
    FromLabelExits.configure(font=text_font)
    FromEntryExits.configure(font=text_font)
    ActualDateButtonExits.configure(font=button_font)
    ExportButtonExits.configure(font=button_font)
    ExportLastDayExits.configure(font=button_font)
    ExportLastMonthExits.configure(font=button_font)
    LabelForces.config(font=title_font)
    ToLabelForces.configure(font=text_font)
    ToEntryForces.configure(font=text_font)
    FromLabelForces.configure(font=text_font)
    FromEntryForces.configure(font=text_font)
    ActualDateButtonForces.configure(font=button_font)
    ExportButtonForces.configure(font=button_font)
    ExportLastDayForces.configure(font=button_font)
    ExportLastMonthForces.configure(font=button_font)
    LabelFaults.config(font=title_font)
    ToLabelFaults.configure(font=text_font)
    ToEntryFaults.configure(font=text_font)
    FromLabelFaults.configure(font=text_font)
    FromEntryFaults.configure(font=text_font)
    ActualDateButtonFaults.configure(font=button_font)
    ExportButtonFaults.configure(font=button_font)
    ExportLastDayFaults.configure(font=button_font)
    ExportLastMonthFaults.configure(font=button_font)
    LabelHabs.config(font=title_font)
    ToLabelHabs.configure(font=text_font)
    ToEntryHabs.configure(font=text_font)
    FromLabelHabs.configure(font=text_font)
    FromEntryHabs.configure(font=text_font)
    ActualDateButtonHabs.configure(font=button_font)
    ExportButtonHabs.configure(font=button_font)
    ExportLastDayHabs.configure(font=button_font)
    ExportLastMonthHabs.configure(font=button_font)
    TableComboxHabs.configure(font=combox_font)
    TableHabs.config(font=text_font)

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

separatorx = (int(ScreenWidth // 50) - int(ScreenHeight // 60))

adjust_size(ScreenWidth, ScreenHeight)

root.mainloop()