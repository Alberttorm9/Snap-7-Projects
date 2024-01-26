import pandas as pd
import os
import pyodbc
from datetime import datetime
import re
import openpyxl
from openpyxl.styles import Font

tabla = "Reportes"
server = 'LAPTOP-3UQV2BFJ\\SQLEXPRESS' 
database = 'Motel_Panamá' 
conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';Trusted_Connection=yes')
cursor = conn.cursor()
query = f"SELECT Reportes_Texto_Reporte, Reportes_Hora_Reporte, Reportes_Numero_Reporte, Reportes_Hora_Real_Reporte, Reportes_Numero_Real_Reporte FROM {tabla} WHERE Time_Stamp BETWEEN '2024-01-24' AND '2024-01-27'"
carpeta_exports = 'Exports\Reportes'
encabezados = ['Texto Del Reporte', 'Hora Del Reporte', 'Numero Del Reporte', 'Hora Real Del Reporte', 'Numero Real Del Reporte']
print(query + "\n\n")
cursor.execute(query)
rows = cursor.fetchall()
rows = [(re.sub(r"[\(\)]", "", item) if isinstance(item, str) else item) for row in rows for item in row]


if not os.path.exists(carpeta_exports):
    os.makedirs(carpeta_exports)


wb = openpyxl.Workbook()
sheet = wb.active

# Agregar encabezados en negrita
for col, encabezado in enumerate(encabezados, 1):
    sheet.cell(row=1, column=col, value=encabezado)
    sheet.cell(row=1, column=col).font = Font(bold=True)

x=0
y=1
for col, rowData in enumerate(rows, 1):
    sheet.cell(row=x+2, column=y, value=rowData)
    y=y+1
    if col==5:
        x=x+1
        y=1
    

# Guardar el archivo
wb.save('export.xlsx')