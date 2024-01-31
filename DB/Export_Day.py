import customtkinter as ctk
import pyodbc
import os
import re
import sys
import configparser
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font

#####################################################################################################################################################

#Configparser
config = configparser.ConfigParser()
config.read(os.path.abspath("Config2.ini"))

#####################################################################################################################################################

# Conexión a la base de datos
server = str(config["DB"]["DBSERVER"]) 
database = str(config["DB"]["DBNAME"]) 
conn = pyodbc.connect('DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';Trusted_Connection=yes')

#####################################################################################################################################################

rutaErrores = f"{config['RUTA']['ERRORES']}/Errors Habs {datetime.now().strftime('%Y-%m-%d')}"

def export_Habs():
    cursor = conn.cursor()
    if os.path.exists(f"{rutaErrores}"):
        os.remove(f"{rutaErrores}")
    for i in range(int(config["HABS"]["NUMERO_HABITACIONES"])) :
        try:
            tabla = f'Tiempo_Limpiando_Habitacion_{i+1}'
            NumeroHab = f'Hab_{i}_Tiempo_Limpiando'
            query = f"SELECT Time_Stamp, {NumeroHab} FROM {tabla} WHERE Time_Stamp >= '{datetime.now().strftime('%Y-%m-%d')} 00:00:00.000000' AND Time_Stamp <= '{datetime.now().strftime('%Y-%m-%d')} 23:59:59.000000'"
            carpeta_exports = str(config["RUTA"]["HABITACIONES"])
            encabezados = ['Hora Terminada', f'Tiempo Limpiando Habitación {i+1}']
            CantidadValores = 2
            cursor.execute(query)
            rows = cursor.fetchall()
            rows = [(re.sub(r"[\(\)]", "", item) if isinstance(item, str) else item) for row in rows for item in row]
            if not os.path.exists(carpeta_exports):
                os.makedirs(carpeta_exports)     
            exportar_excel(rows, encabezados, CantidadValores, carpeta_exports, tabla)
        except Exception as e:
            if not os.path.exists(rutaErrores):
                os.makedirs(rutaErrores)
            if re.search(r"42S02", str(e)):
                with open(rutaErrores, 'w') as archivo:   
                    archivo.write(f'Error: No existe la tabla de la habitacion {i+1}\n')

def export_Reports():
    cursor = conn.cursor()
    if os.path.exists(f"{rutaErrores}/Errors Reports {datetime.now().strftime('%Y-%m-%d')}"):
        os.remove(f"{rutaErrores}/Errors Reports {datetime.now().strftime('%Y-%m-%d')}")
    try:
        tabla = f'Reportes'
        query = f"SELECT Reportes_Texto_Reporte, Reportes_Hora_Reporte, Reportes_Numero_Reporte, Reportes_Hora_Real_Reporte, Reportes_Numero_Real_Reporte FROM {tabla} WHERE Time_Stamp >= '{datetime.now().strftime('%Y-%m-%d')} 00:00:00.000000' AND Time_Stamp <= '{datetime.now().strftime('%Y-%m-%d')} 23:59:59.000000'"
        carpeta_exports = str(config["RUTA"]["REPORTES"])
        encabezados = ('Texto Del Reporte', 'Hora Del Reporte', 'Numero Del Reporte', 'Hora Real Del Reporte', 'Numero Real Del Reporte')
        CantidadValores = 5
        cursor.execute(query)
        rows = cursor.fetchall()
        rows = [(re.sub(r"[\(\)]", "", item) if isinstance(item, str) else item) for row in rows for item in row]
        if not os.path.exists(carpeta_exports):
            os.makedirs(carpeta_exports)     
        exportar_excel(rows, encabezados, CantidadValores, carpeta_exports, tabla)
    except Exception as e:
        if re.search(r"42S02", str(e)):
            with open(f"{rutaErrores}/Errors Reports {datetime.now().strftime('%Y-%m-%d')}", 'w') as archivo:   
                archivo.write(f'Error: No existe la tabla de Reportes\n')

def export_Exits():
    cursor = conn.cursor()
    if os.path.exists(f"{rutaErrores}/Errors Exits {datetime.now().strftime('%Y-%m-%d')}"):
        os.remove(f"{rutaErrores}/Errors Exits {datetime.now().strftime('%Y-%m-%d')}")
    try:
        tabla = f'Salida_De_Programa'
        query = f"SELECT Now_Local, sys_On_Off FROM {tabla} WHERE Time_Stamp >= '{datetime.now().strftime('%Y-%m-%d')} 00:00:00.000000' AND Time_Stamp <= '{datetime.now().strftime('%Y-%m-%d')} 23:59:59.000000'"
        carpeta_exports = str(config["RUTA"]["SALIDAS"])
        encabezados = ['Hora De Accion', 'Tipo De Accion']
        CantidadValores = 2
        cursor.execute(query)
        rows = cursor.fetchall()
        rows = [(re.sub(r"[\(\)]", "", item) if isinstance(item, str) else item) for row in rows for item in row]
        if not os.path.exists(carpeta_exports):
            os.makedirs(carpeta_exports)     
        exportar_excel(rows, encabezados, CantidadValores, carpeta_exports, tabla)
    except Exception as e:
        if re.search(r"42S02", str(e)):
            with open(f"{rutaErrores}/Errors Exits {datetime.now().strftime('%Y-%m-%d')}", 'w') as archivo:   
                archivo.write(f'Error: No existe la tabla de Salidas De Sistema\n')

def exportar_excel(rows, encabezados,CantidadValores, carpeta_exports, tabla):
    wb = openpyxl.Workbook()
    sheet = wb.active
    #Encabezado
    for col, encabezado in enumerate(encabezados, 1):
        sheet.cell(row=1, column=col, value=encabezado)
        sheet.cell(row=1, column=col).font = Font(bold=True)
    #Lineas
    x=0
    y=0
    for col, rowData in enumerate(rows, 1):
        y=y+1
        sheet.cell(row=x+2, column=y, value=rowData)
        if col % CantidadValores == 0:
            x=x+1
            y=0

    wb.save(f'{carpeta_exports}\{tabla} {datetime.now().strftime("%d-%m-%y")}.xlsx')

def export_elements():
    export_Habs()
    export_Reports()
    export_Exits()

export_elements()            